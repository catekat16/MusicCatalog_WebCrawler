from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.action_chains import ActionChains
import pandas as pd
import time
from collections import defaultdict

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


def check_master(song_title, index_check=False):
    substrings = ['Master','Mast', 'ssh', 'bmi']
    if not index_check:
        return any([substring.lower() in song_title.lower() for substring in substrings])
    else:
        return max([song_title.lower().find(substring.lower()) for substring in substrings])

def format_copynum(song_title):
    string1 = song_title
    string1 = string1.split("-")
    string_list = []
    for string in string1:
        string_list.append(string.split(" "))

    flatten_list = sum(string_list, [])

    length = 0
    for i in range(len(flatten_list)):
        length += len(flatten_list[i])

    song_name = ""

    for index in range(len(flatten_list)):
        if index == 0:
            song_name += flatten_list[index]
            for j in range(12-length):
                song_name += "0"
            continue
        song_name += flatten_list[index]
        
    return song_name

def create_check_dict(testfile_name): 
    require_cols = [2, 3]

    dataframe2 = pd.read_excel(testfile_name + '.xlsx', engine = 'openpyxl', usecols = require_cols, skiprows = 2, header = None).head(-142)
    dataframe2 = dataframe2.fillna("")
    
    song_titles = []
    check_dict = defaultdict(str)

    for index, row in dataframe2.iterrows():
        copy_num = format_copynum(row[3])
        if copy_num[:2] in ["PA", "PAu", "PAU", "SR"]: 
            #check_dict[(index+1,(row[2]).split("-")[1])] = copy_num
            check_dict[(index+1,row[2])] = copy_num
    return check_dict
    
def create_output(copyright_dict, outputfile_name, accuracy_check=False):
    wb =  Workbook()
    ws =  wb.active
    headings = ["#","Song","Registration #"]
    ws.append(headings)
    
    if accuracy_check:
        check_dict = create_check_dict("test_Song_Registration Information")
        match = 0
        #print("check_dict:", check_dict)
    
    num_recorded = 0
    for key in copyright_dict.keys():
        #char = get_column_letter(col)
        # easy mvp but need to store original row number and do something like ws[char + row_iter]
        row_num = key[0]
        print("\nrow_num: ", row_num)
        song_title = key[1]
        # want the original song title (in the case of MASTER version of 20th Century Boy and regular 20th Century Boy, for example)
        copy_num = copyright_dict[key]
        #print(song_title, end="")
        #print(copy_num)
        #print("check_dict[row_num, song_title]: ", check_dict[(row_num,song_title)])
        print("song_title: ", song_title)
        if accuracy_check:
            if check_dict[(row_num,song_title)] == " ":
                print("non PA/SR number entry -- won't be factored into total accuracy count")
                num_recorded -= 1
            elif check_dict[(row_num,song_title)][:2] not in ["PA", "SR"]:
                num_recorded -= 1
            else:
                print("copy_num: ", copy_num)
                print("Manual entry (check_dict[row_num, song_title]) is: ", check_dict[(row_num,song_title)])
                if check_dict[(row_num,song_title)].lower() == copy_num.lower():
                    print("match!")
                    match += 1
        ws.append([row_num, song_title, copy_num])
        if copy_num != " ":
            num_recorded += 1

    wb.save(outputfile_name + '.xlsx')
    if accuracy_check:
        accuracy = match/num_recorded * 100
    total_num_items = len(copyright_dict)
    print("\nCheck out the output file!")
    print("Accounting for", num_recorded, "songs out of", len(copyright_dict))
    print("Match against manually inputted file is: ", accuracy, "%")
    print(match, " matches out of ", num_recorded, " registration numbers recorded")
    return [num_recorded, total_num_items, accuracy, match]

def run_main(input_file):

    start_time = time.time()
    require_cols = [2, 2]
    dataframe1 = pd.read_excel(input_file, engine = 'openpyxl', usecols = require_cols, skiprows = 2, header = None).head(-142) 
    #dataframe1 = pd.read_excel('/Users/xiaoyanyang/Desktop/MusicCatalog_WebCrawler/input_Song_Titles.xlsx', engine = 'openpyxl', usecols = require_cols, skiprows = 2, header = None).head(-142)
    copyright_dict = defaultdict(str)
    #song_titles1 = []
    song_titles = []
    #breaks_code = ['HUMBLE AND KIND (TIM MCGRAW MA', 'HOLIDAY (SSH)', "CAN'T YOU SEE (U.S. ONLY AS OF)", "HIGHWAY DON'T CARE (TIM MCGRAW", '25 OR 6 TO 4 (GOARMY REMIX)', 'DANCE HALL DAYS (WANG CHUNG RE', 'SHOTGUN RIDER (TIM MCGRAW MAST', 'SHOTGUN RIDER (TIM MCGRAW MAST', 'BAREFOOT BLUE JEAN NIGHT (BMI']
    #breaks_code = ['HUMBLE AND KIND (TIM MCGRAW MA', "HIGHWAY DON'T CARE (TIM MCGRAW", 'HOLIDAY (SSH)', 'BODY', 'LET MY LOVE OPEN THE DOOR', "WON'T GET FOOLED AGAIN", 'LOOK WHAT YOU MADE ME DO', '25 OR 6 TO 4 (GOARMY REMIX)', 'DANCE HALL DAYS (WANG CHUNG RE', 'SHOTGUN RIDER (TIM MCGRAW MAST']
    ignore = ['LET MY LOVE OPEN THE DOOR', "WON'T GET FOOLED AGAIN", 'LOOK WHAT YOU MADE ME DO']

    for index, row in dataframe1.iterrows():
        if row[2].split("-")[1] in ignore:
        #if len(row[2].split("-")[1].split("(")[0]) == 1: # single word case
            continue
        else:
            #if "(" in row[2]:
            #    first_half = row[2].split("(")[0]
            #    song_titles.append((first_half.split("-"))[1])
            #else:
            #row_item = row[2].replace("(", "")
            #    song_titles.append((row[2]).split("-")[1])
            song_titles.append((index+1,row[2]))

    PATH = "/Users/xiaoyanyang/Desktop/chromedriver"
    driver = webdriver.Chrome(PATH)
    driver.get("https://cocatalog.loc.gov/cgi-bin/Pwebrecon.cgi?DB=local&PAGE=first")
    
    #song_titles = ["20th Century Boy", "Burn Out", "Glorious Domination (Master)", "More Girls Like You", "Stars in the night"]
    #song_titles = ["Glorious Domination (Master)"]
    #song_titles = ["CAN'T YOU SEE (U.S. ONLY AS OF)"]
    #song_title = ["HUMBLE AND KIND (TIM MCGRAW MA"]

    #if "(" in song_title[0]:
    #    first_half = row[2].split("(")[0]
    #    song_titles.append((first_half.split("-"))[1])

    #song_title = ["202775-HUMBLE AND KIND (TIM MCGRAW MA"]
    #song_titles.append((6,song_title[0]))
    
    ######## TEST CASE #########
    
    #song_title = ["19819-20th Century Boy - Master"]
    #song_titles.append((12,song_title[0]))
    #if check_master(song_title[0]):
    #    SR_master = True
    #print(song_titles)
    #print(SR_master)
    
    ############################
    
    registered_count = 0
    for i in range(len(song_titles)):
        SR_master = False
        song_index = song_titles[i][0]
        song_title = song_titles[i][1]
        if check_master(song_title):
            SR_master = True
        # search_term can be song title + catalog (try different search terms)
        
        search_term = song_title.split("-")[1]
        #print("\nsearch_term after splitting for -: ", search_term, "\n")
        if check_master(search_term): # gets rid of the phrase master, ssh, etc.
            end_index = check_master(search_term, True)
            search_term = search_term[:end_index-1]
        if search_term[-1] == " ":
            search_term = search_term[:-1]

        search_term = search_term.split("(")[0]
            
        if len(search_term.split(" ")) == 1:
            continue
        #print("\nsearch_term: ", search_term, "\n")

        search = driver.find_element_by_name("Search_Arg")
        search.clear()
        search.send_keys(search_term + "\n")

        driver.implicitly_wait(0.5)
        
        possibilities_list = []
        
        try:
            table = driver.find_elements_by_xpath("//form/table[2]/tbody/tr")
        except:
            table = driver.find_elements_by_xpath("//form/table[2]")
            
        #try:
        num_cols = len(table[1].text.split(":"))
        #except:
            #search = driver.find_element_by_name("Search_Arg")
            #search.clear()
            #search_term = search_term.split("(")[0]
            #search.send_keys(search_term + "\n")
            #print("\nnew search_term: ", search_term, "\n")

        #print("number of cols is", num_cols)
        if num_cols == 2: # details page of single search result case
            row_item = table[2].text
            #print("row item is", row_item)
            try: 
                copy_num = row_item.split(":")[-1].split("/")[-2]
            except:
                # /html/body/form[1]/table[2]/tbody/tr[3]/th
                # need to figure out how to access the column over (while keeping in mind that registration number can show up in different rows)
                row_index = driver.find_elements_by_xpath("//*[contains(text(), 'Registration Number')]").index
                #print(row_index)
            possibilities_list.append(copy_num) 
            
        else: # more than one search result case
            if not i:
                sel = Select(driver.find_element_by_xpath("//select[@name='CNT']"))
                sel.select_by_value("100")
                driver.implicitly_wait(0.5)
                submit_button = driver.find_element_by_xpath("//form/center[2]/form/table/tbody/tr[2]/td/div/div/table/tbody/tr/td[2]/div/input[2]")
                submit = ActionChains(driver)
                submit.click(submit_button)
                submit.perform()
            
            try:
                table = driver.find_elements_by_xpath("//form/table[2]/tbody/tr")
            except:
                table = driver.find_elements_by_xpath("//form/table[2]")
            num_rows = len(table)
            #print("\nnum_rows: ", num_rows, "\n")
            for i in range(1,num_rows):
                row_item = table[i].text #
                copy_num = row_item.split(".")[-1].split(" ")[-2]
                #print("\n\n\n\ncopy_num: " + copy_num + "\n\n\n\n")
                song_info = row_item
                if (search_term.lower() in search_term.lower()):
                    date = row_item.split(".")[-1].split(" ")[-1]
                    if SR_master:
                        #print(SR_master)
                        if copy_num[:2] == "SR":
                            #print("\n\n\n\n------------SR case--------------\n\n\n\n")
                            # possibilities_list.append((copy_num, date)) 
                            # instead of appending date, in better MVP would click into details page 
                            #     and see whether catalog name is listed in authorship on application and append type of work 
                            possibilities_list.append(copy_num)

                    elif copy_num[:2] == "PA" and copy_num[:3] != "PAu": 
                        possibilities_list.append(copy_num) 

        #{(song name, row num):copyright #} dictionary
        # in the excel file row num offset by +2
        if not possibilities_list:
            copyright_dict[(song_index, song_title)] = " "
        else:
            registered_count += 1
            copyright_dict[(song_index, song_title)] = possibilities_list[0]
            #print("registered song")

        print("Number of songs registered: ", registered_count)
        print(song_title)
        #print("\npossibilities_list: ", possibilities_list, "\n")
        
        try:
            search_toggle = driver.find_element_by_xpath("//center[2]/font/a[2]/img")
        except:
            search_toggle = driver.find_element_by_xpath("//center[2]/font/img")
        finally:
            back_tosearch = ActionChains(driver)
            back_tosearch.click(search_toggle)
            back_tosearch.perform()
            
    driver.close()
    #print(copyright_dict) 
    #print("copyright_dict: ", copyright_dict)
    
    output_info = create_output(copyright_dict, "Registration Numbers", True)
    seconds_elapsed = time.time() - start_time
    print("Total time elapsed (in minutes): ", seconds_elapsed/60)
    average_seconds = seconds_elapsed/len(copyright_dict)
    print("Average time taken for each song (in seconds): ", average_seconds)
    return [(seconds_elapsed, average_seconds),output_info]
# next steps: 
# yup --> - paste to excel file (just take first element of list)
# - use more complete information of excel sheet (catalog name, artist name, date)
# - figure out how to seek registration number dynamically within details page of song
# turn into object-oriented so that things like copyright_dict are global
# build a web interface, start the process after uploading input file and pressing button 
# yup --> fix 20th century boy Master version (don't split at the beginning right away on "(" if detect Master)
# if single word keyword song (Holiday, Body), search on more information
# look up web crawler deepening search techniques / algorithms